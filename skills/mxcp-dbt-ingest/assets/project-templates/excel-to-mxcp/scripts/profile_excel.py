#!/usr/bin/env python3
"""
Excel Data Profiler — produces a comprehensive markdown report.

Usage:
    python scripts/profile_excel.py <excel_file> [--output data-profile-report.md]

Capabilities:
- Auto-detects header row (handles title rows, multi-line headers)
- Per-column: type inference with mixed-type warnings, null/empty-string tracking
- Locale-aware number parsing (European decimals, currency symbols)
- Merged cell region mapping with date-serial-number awareness
- Hidden sheet detection and formula-cell warnings
- Transposition detection
- Cross-sheet FK candidate detection with cardinality guards
- Calculated column detection with numpy.isclose tolerance
- Chunked reading for 50MB+ files (openpyxl read-only mode)
"""

import argparse
import hashlib
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

# Import shared utilities from canonical module (single source of truth)
sys.path.insert(0, str(Path(__file__).parent))
from canonical import normalize_column_name, strip_currency_pct, try_european_decimal, convert_excel_date


# --- File metadata ---

def get_file_metadata(filepath: Path) -> dict:
    """Get basic file metadata."""
    stat = filepath.stat()
    hasher = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            hasher.update(chunk)
    file_hash = hasher.hexdigest()
    return {
        "path": str(filepath.resolve()),
        "size_bytes": stat.st_size,
        "size_mb": round(stat.st_size / (1024 * 1024), 2),
        "md5": file_hash,
        "profiled_at": datetime.now().isoformat(),
    }


# --- Sheet-level detection ---

def detect_hidden_sheets(filepath: Path) -> list:
    """Detect hidden and very-hidden sheets."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    hidden = []
    for name in wb.sheetnames:
        state = wb[name].sheet_state
        if state != "visible":
            hidden.append({"name": name, "state": state})
    wb.close()
    return hidden


def detect_named_ranges(filepath: Path) -> list:
    """Detect Excel named ranges (defined names) that may reference data regions."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    named_ranges = []
    # openpyxl >= 3.1: wb.defined_names is a DefinedNameDict (iterable over DefinedName objects)
    # openpyxl < 3.1: wb.defined_names.definedName was the list attribute
    names_iter = wb.defined_names.values() if hasattr(wb.defined_names, 'values') else wb.defined_names
    for defn in names_iter:
        try:
            destinations = list(defn.destinations)
            for sheet_title, cell_range in destinations:
                named_ranges.append({
                    "name": defn.name,
                    "sheet": sheet_title,
                    "range": cell_range,
                    "hidden": getattr(defn, 'hidden', False) or False,
                })
        except Exception:
            # Some named ranges (e.g., print areas) may not resolve cleanly
            named_ranges.append({
                "name": defn.name,
                "sheet": "(unresolved)",
                "range": str(getattr(defn, 'attr_text', defn.value)),
                "hidden": getattr(defn, 'hidden', False) or False,
            })
    wb.close()
    return named_ranges


def detect_mojibake(series: pd.Series, sample_size: int = 200) -> dict | None:
    """Detect mojibake (encoding corruption) patterns in string columns.

    Common patterns: Ã¡ instead of á, â€™ instead of ', Ã© instead of é.
    Returns None if no issues detected, otherwise a dict with details.
    """
    MOJIBAKE_PATTERNS = [
        r'Ã[¡¤¥¨©ª«¬®¯°±²³´µ¶·¸¹º»¼½¾¿]',  # UTF-8 misread as Latin-1
        r'â€[™˜œ""\u009c\u009d]',                  # Smart quotes garbled
        r'Ã‰|Ã¨|Ã©|Ã¼|Ã¶|Ã¤',                    # Accented chars garbled
        r'[\x80-\x9f]',                              # Windows-1252 control chars
    ]
    str_vals = series.dropna().astype(str)
    if len(str_vals) == 0:
        return None
    sample = str_vals.head(sample_size)
    hits = 0
    examples = []
    for val in sample:
        for pattern in MOJIBAKE_PATTERNS:
            if re.search(pattern, val):
                hits += 1
                if len(examples) < 3:
                    examples.append(val[:80])
                break
    if hits > 0:
        return {
            "affected_count": hits,
            "sample_size": len(sample),
            "rate": round(hits / len(sample), 4),
            "examples": examples,
        }
    return None


def detect_formula_columns(filepath: Path, sheet_name: str, header_row: int = 0) -> list:
    """Detect columns with formulas that may have no cached values."""
    warnings = []
    # Convert 0-indexed pandas header_row to 1-indexed openpyxl row
    header_row_1based = (header_row + 1) if isinstance(header_row, int) else (header_row[-1] + 1) if isinstance(header_row, list) else 1
    try:
        wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
        wb_values = openpyxl.load_workbook(filepath, data_only=True)
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]

        # Check first 20 data rows for formula cells with None cached values
        formula_cols = defaultdict(lambda: {"formula_count": 0, "none_count": 0})
        data_start = header_row_1based + 1
        for row_idx in range(data_start, min(data_start + 20, ws_f.max_row + 1)):
            for col_idx in range(1, ws_f.max_column + 1):
                cell_f = ws_f.cell(row_idx, col_idx)
                cell_v = ws_v.cell(row_idx, col_idx)
                if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    col_name = ws_f.cell(header_row_1based, col_idx).value or f"Col{col_idx}"
                    formula_cols[col_name]["formula_count"] += 1
                    if cell_v.value is None:
                        formula_cols[col_name]["none_count"] += 1

        for col_name, counts in formula_cols.items():
            if counts["none_count"] > counts["formula_count"] * 0.5:
                warnings.append(
                    f"Column '{col_name}' contains formulas with no cached values "
                    f"({counts['none_count']}/{counts['formula_count']} are None). "
                    f"The file may need to be opened and saved in Excel first."
                )
        wb_formulas.close()
        wb_values.close()
    except Exception:
        pass  # Non-critical, skip if can't check
    return warnings


def detect_header_row(filepath: Path, sheet_name: str) -> int | list:
    """Auto-detect which row(s) are the actual header (0-indexed for pandas).

    Returns:
        int — single header row index (most common case)
        list[int] — multi-row header indices (e.g., [0, 1] for 2-row headers)

    Heuristics:
    1. Scan first 10 rows, find the row with the most unique non-empty text values.
    2. Check if the row ABOVE has fewer text values that look like category groupings
       (merged or sparse text above a detail-column row). If so, return both as
       multi-row header for pandas multi-level column support.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True))
    wb.close()

    if not rows:
        return 0

    best_row = 0
    best_score = 0
    row_scores = []

    for idx, row in enumerate(rows):
        vals = [v for v in row if v is not None and str(v).strip()]
        if not vals:
            row_scores.append(0)
            continue
        text_count = sum(1 for v in vals if isinstance(v, str) and not v.replace('.','').replace('-','').isdigit())
        unique_count = len(set(str(v) for v in vals))
        coverage = len(vals) / max(len(row), 1)
        score = text_count * unique_count * coverage
        row_scores.append(score)
        if score > best_score:
            best_score = score
            best_row = idx

    # Multi-row header detection: check if row above the best row looks like
    # category groupings (fewer unique values spanning groups above detail columns)
    if best_row > 0:
        above_row = rows[best_row - 1]
        above_vals = [v for v in above_row if v is not None and str(v).strip()]
        above_text = sum(1 for v in above_vals if isinstance(v, str))

        if above_text >= 2:
            above_unique = len(set(str(v) for v in above_vals if v is not None))
            best_vals = [v for v in rows[best_row] if v is not None and str(v).strip()]
            # Multi-row if above has text but fewer values than detail row
            if 2 <= len(above_vals) < len(best_vals) and above_unique < len(above_vals) * 0.8:
                return [best_row - 1, best_row]

    return best_row


def detect_merged_cells(filepath: Path) -> dict:
    """Map all merged cell regions per sheet using openpyxl."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    merged = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        regions = []
        for merge_range in ws.merged_cells.ranges:
            raw_value = ws.cell(merge_range.min_row, merge_range.min_col).value
            # Handle date serial numbers from openpyxl
            if isinstance(raw_value, (int, float)) and 1 < raw_value < 200000:
                try:
                    as_date = datetime(1899, 12, 30) + timedelta(days=int(raw_value))
                    display = f"{raw_value} (possible date: {as_date.strftime('%Y-%m-%d')})"
                except Exception:
                    display = str(raw_value)
            else:
                display = str(raw_value) if raw_value is not None else "(empty)"
            regions.append({
                "range": str(merge_range),
                "rows": f"{merge_range.min_row}-{merge_range.max_row}",
                "cols": f"{merge_range.min_col}-{merge_range.max_col}",
                "value": display,
            })
        if regions:
            merged[sheet_name] = regions
    wb.close()
    return merged


def detect_transposition(df: pd.DataFrame) -> dict:
    """Detect if a table is likely transposed."""
    if df.empty or len(df.columns) < 3:
        return {"likely_transposed": False, "reason": "too few columns"}

    first_col = df.iloc[:, 0]
    other_cols = df.iloc[:, 1:]
    first_is_text = first_col.apply(lambda x: isinstance(x, str)).mean() > 0.7
    others_numeric = other_cols.apply(pd.to_numeric, errors="coerce").notna().mean().mean() > 0.7

    if first_is_text and others_numeric:
        headers = [str(c) for c in other_cols.columns]
        date_like = sum(1 for h in headers if any(m in h.lower() for m in
            ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
             "q1", "q2", "q3", "q4", "2020", "2021", "2022", "2023", "2024", "2025", "2026"])) / max(len(headers), 1)
        numeric_headers = sum(1 for h in headers if h.replace(".", "").replace("-", "").isdigit()) / max(len(headers), 1)
        if date_like > 0.5 or numeric_headers > 0.5:
            return {
                "likely_transposed": True,
                "reason": f"text first column + numeric data + {'date-like' if date_like > 0.5 else 'numeric'} headers",
                "suggestion": "Unpivot: use pandas.melt() with first column as id_vars",
            }

    return {"likely_transposed": False, "reason": "standard orientation"}


# --- Column profiling ---

def profile_column(series: pd.Series) -> dict:
    """Profile a single column with mixed-type detection and empty-string tracking."""
    total = len(series)
    null_count = int(series.isna().sum())
    non_null = series.dropna()

    # Track empty strings and whitespace-only separately from NULLs
    str_vals = non_null.astype(str)
    empty_string_count = int((str_vals == "").sum())
    whitespace_only_count = int((str_vals.str.strip() == "").sum()) - empty_string_count

    profile = {
        "null_count": null_count,
        "null_rate": round(null_count / max(total, 1), 4),
        "empty_string_count": empty_string_count,
        "whitespace_only_count": whitespace_only_count,
        "unique_count": int(non_null.nunique()),
        "unique_rate": round(non_null.nunique() / max(len(non_null), 1), 4),
    }

    if len(non_null) == 0:
        profile["inferred_type"] = "empty"
        return profile

    # --- Try numeric (with currency/locale stripping) ---
    # First pass: standard numeric
    numeric = pd.to_numeric(non_null, errors="coerce")
    numeric_rate = numeric.notna().mean()

    # Second pass: strip currency/percentage and retry failed values
    if numeric_rate < 0.9:
        cleaned = non_null.apply(lambda x: strip_currency_pct(str(x)) if isinstance(x, str) else x)
        numeric_cleaned = pd.to_numeric(cleaned, errors="coerce")
        # Third pass: try European decimal format on remaining failures
        still_nan = numeric_cleaned.isna() & non_null.notna()
        if still_nan.sum() > 0:
            euro_parsed = non_null[still_nan].apply(lambda x: try_european_decimal(str(x)))
            numeric_cleaned[still_nan] = euro_parsed
        cleaned_rate = numeric_cleaned.notna().mean()
        if cleaned_rate > numeric_rate:
            numeric = numeric_cleaned
            numeric_rate = cleaned_rate
            if cleaned_rate > 0.9:
                profile["format_detected"] = "currency/locale-formatted numbers"

    if numeric_rate > 0.9:
        clean_numeric = numeric.dropna()
        # Safe integer check (avoid overflow for large values)
        try:
            is_integer = (clean_numeric == clean_numeric.round(0)).all() and clean_numeric.abs().max() < 2**53
        except (OverflowError, ValueError):
            is_integer = False
        profile["inferred_type"] = "integer" if is_integer else "number"
        profile["min"] = float(clean_numeric.min())
        profile["max"] = float(clean_numeric.max())
        profile["mean"] = round(float(clean_numeric.mean()), 4)
        profile["samples"] = [str(x) for x in non_null.head(3).tolist()]
        # Warn about mixed types if coercion dropped some values
        if 0.5 < numeric_rate < 0.95:
            failed_count = int((~numeric.notna() & non_null.notna()).sum())
            profile["mixed_type_warning"] = f"{failed_count} values failed numeric coercion"
            profile["coercion_failures"] = [str(x) for x in non_null[~numeric.notna() & non_null.notna()].head(5).tolist()]
        return profile

    # --- Try datetime (raised threshold + date-range plausibility guard) ---
    try:
        dates = pd.to_datetime(non_null, errors="coerce", format="mixed")
        date_rate = dates.notna().mean()
        if date_rate > 0.95:  # Raised from 0.8 to prevent integer-as-date misclassification
            clean_dates = dates.dropna()
            # Plausibility guard: reject if dates are outside 1900-2100
            min_date = clean_dates.min()
            max_date = clean_dates.max()
            if min_date.year >= 1900 and max_date.year <= 2100:
                profile["inferred_type"] = "datetime"
                profile["min"] = str(min_date)
                profile["max"] = str(max_date)
                profile["samples"] = [str(x) for x in non_null.head(3).tolist()]
                return profile
    except Exception:
        pass

    # --- Try boolean ---
    bool_values = {"true", "false", "yes", "no", "y", "n", "1", "0"}
    if non_null.astype(str).str.lower().isin(bool_values).mean() > 0.9:
        profile["inferred_type"] = "boolean"
        profile["values"] = sorted(non_null.astype(str).str.lower().unique().tolist())
        return profile

    # --- Default: string ---
    profile["inferred_type"] = "string"
    profile["avg_length"] = round(non_null.astype(str).str.len().mean(), 1)
    profile["max_length"] = int(non_null.astype(str).str.len().max())
    profile["samples"] = [str(x) for x in non_null.head(3).tolist()]

    if profile["unique_count"] <= 20 and profile["unique_rate"] < 0.1:
        profile["possible_enum"] = True
        profile["values"] = sorted(non_null.astype(str).unique().tolist()[:20])

    return profile


# --- Relationship detection ---

def detect_fk_candidates(sheets_data: dict) -> list:
    """Detect cross-sheet foreign key candidates with cardinality and name guards."""
    column_values = {}
    column_cardinality = {}
    for sheet_name, df in sheets_data.items():
        for col in df.columns:
            non_null = df[col].dropna()
            uniq = non_null.nunique()
            if len(non_null) > 0 and uniq > 1:
                key = (sheet_name, col)
                column_values[key] = set(non_null.astype(str).unique())
                column_cardinality[key] = uniq

    candidates = []
    keys = list(column_values.keys())
    for i, (sheet_a, col_a) in enumerate(keys):
        for sheet_b, col_b in keys[i + 1:]:
            if sheet_a == sheet_b:
                continue

            # Cardinality guard: skip if BOTH sides have very low cardinality (<20 unique)
            # Low-cardinality matches (status codes, booleans) are almost never FKs
            if column_cardinality[(sheet_a, col_a)] < 20 and column_cardinality[(sheet_b, col_b)] < 20:
                continue

            vals_a = column_values[(sheet_a, col_a)]
            vals_b = column_values[(sheet_b, col_b)]
            if len(vals_a) == 0 or len(vals_b) == 0:
                continue

            overlap = vals_a & vals_b
            match_rate_a = len(overlap) / len(vals_a)
            match_rate_b = len(overlap) / len(vals_b)

            # Column name similarity boost (shared suffix like "_id" or same normalized name)
            norm_a = normalize_column_name(col_a)
            norm_b = normalize_column_name(col_b)
            name_similar = norm_a == norm_b or norm_a.endswith("_id") and norm_b.endswith("_id")
            threshold = 0.7 if name_similar else 0.8

            if match_rate_a > threshold or match_rate_b > threshold:
                # Determine FK direction: many-side references one-side
                if len(vals_b) <= len(vals_a) and match_rate_a > threshold:
                    candidates.append({
                        "from_sheet": sheet_a, "from_column": col_a,
                        "to_sheet": sheet_b, "to_column": col_b,
                        "match_rate": round(match_rate_a, 4),
                        "orphan_count": len(vals_a) - len(overlap),
                        "direction": f"{sheet_a}.{col_a} -> {sheet_b}.{col_b}",
                        "name_similarity": name_similar,
                    })
                elif len(vals_a) <= len(vals_b) and match_rate_b > threshold:
                    candidates.append({
                        "from_sheet": sheet_b, "from_column": col_b,
                        "to_sheet": sheet_a, "to_column": col_a,
                        "match_rate": round(match_rate_b, 4),
                        "orphan_count": len(vals_b) - len(overlap),
                        "direction": f"{sheet_b}.{col_b} -> {sheet_a}.{col_a}",
                        "name_similarity": name_similar,
                    })

    return candidates


# --- Calculated column detection ---

def detect_calculated_columns(df: pd.DataFrame) -> list:
    """Detect columns calculated from other columns using numpy.isclose."""
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    if len(numeric_cols) < 2:
        return []

    calculated = []
    seen_targets = set()  # Deduplicate results per target column

    for target_col in numeric_cols:
        target = df[target_col].dropna()
        if len(target) < 5:
            continue

        for i, col_a in enumerate(numeric_cols):
            if col_a == target_col:
                continue
            for col_b in numeric_cols[i + 1:]:
                if col_b == target_col:
                    continue

                shared_idx = target.index.intersection(df[col_a].dropna().index).intersection(df[col_b].dropna().index)
                if len(shared_idx) < 5:
                    continue

                a = df.loc[shared_idx, col_a].values
                b = df.loc[shared_idx, col_b].values
                t = df.loc[shared_idx, target_col].values

                for op_name, result in [("*", a * b), ("+", a + b), ("-", a - b)]:
                    if np.isnan(result).all():
                        continue
                    close = np.isclose(t, result, rtol=1e-2, atol=1e-6, equal_nan=True)
                    match_rate = float(close.mean())
                    if match_rate > 0.95:
                        dedup_key = (target_col, op_name)
                        if dedup_key in seen_targets:
                            continue
                        seen_targets.add(dedup_key)
                        mismatch_count = int((~close).sum())
                        # Collect mismatch examples for diagnosis
                        mismatch_examples = []
                        if mismatch_count > 0:
                            mismatch_idx = np.where(~close)[0][:3]
                            for mi in mismatch_idx:
                                mismatch_examples.append({
                                    "row": int(shared_idx[mi]),
                                    col_a: float(a[mi]), col_b: float(b[mi]),
                                    target_col: float(t[mi]),
                                    "expected": float(result[mi]),
                                    "diff": float(abs(t[mi] - result[mi])),
                                })
                        calculated.append({
                            "column": target_col,
                            "formula": f"{col_a} {op_name} {col_b}",
                            "match_rate": round(match_rate, 4),
                            "mismatches": mismatch_count,
                            "mismatch_examples": mismatch_examples,
                        })

    return calculated


# --- Sheet reading with chunking ---

def read_sheet(filepath: Path, sheet_name: str, file_size_mb: float) -> pd.DataFrame:
    """Read a sheet, using chunked openpyxl for large files (50MB+).

    Handles both single-row and multi-row headers from detect_header_row().
    """
    header_row = detect_header_row(filepath, sheet_name)
    is_multi_header = isinstance(header_row, list)

    if file_size_mb <= 50:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)
        # Flatten multi-level columns to "Category_Detail" format
        if is_multi_header and isinstance(df.columns, pd.MultiIndex):
            df.columns = ['_'.join(str(c) for c in col if str(c) != '').strip('_')
                          for col in df.columns.values]
    else:
        # Chunked reading via openpyxl read-only mode
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows = []
        header = None
        first_data_row = (max(header_row) + 1) if is_multi_header else (header_row + 1)
        header_start = (min(header_row)) if is_multi_header else header_row

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx < header_start:
                continue  # Skip title rows above header
            if is_multi_header and row_idx in header_row:
                if header is None:
                    header = [str(c) if c is not None else "" for c in row]
                else:
                    # Merge with previous header row
                    header = [f"{h}_{str(c)}" if c is not None and str(c).strip() else h
                              for h, c in zip(header, row)]
                continue
            if not is_multi_header and row_idx == header_row:
                header = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(row)]
                continue
            rows.append(row)
        wb.close()
        df = pd.DataFrame(rows, columns=header) if header else pd.DataFrame(rows)

    df = df.dropna(how="all").dropna(axis=1, how="all")

    if is_multi_header:
        print(f"  Note: multi-row header detected at rows {[r+1 for r in header_row]}")
    elif header_row > 0:
        print(f"  Note: header detected at row {header_row + 1} (skipped {header_row} title row(s))")

    return df


# --- Report generation ---

def generate_report(filepath: Path, output: Path) -> None:
    """Generate the full profiling report as markdown."""
    file_size_mb = filepath.stat().st_size / (1024 * 1024)
    if file_size_mb > 50:
        print(f"NOTE: Large file ({file_size_mb:.0f}MB). Using chunked reading mode.")
    print(f"Profiling {filepath}...")

    metadata = get_file_metadata(filepath)
    merged = detect_merged_cells(filepath)
    hidden_sheets = detect_hidden_sheets(filepath)
    named_ranges = detect_named_ranges(filepath)

    # Read all sheets
    excel_file = pd.ExcelFile(filepath)
    sheets_data = {}
    formula_warnings = {}
    for sheet_name in excel_file.sheet_names:
        try:
            df = read_sheet(filepath, sheet_name, file_size_mb)
            sheets_data[sheet_name] = df
            # Check for formula columns with no cached values
            header_row = detect_header_row(filepath, sheet_name)
            fw = detect_formula_columns(filepath, sheet_name, header_row)
            if fw:
                formula_warnings[sheet_name] = fw
        except Exception as e:
            sheets_data[sheet_name] = pd.DataFrame()
            print(f"  Warning: could not read sheet '{sheet_name}': {e}")

    fk_candidates = detect_fk_candidates(sheets_data)

    # Build report
    lines = []
    lines.append("# Data Profile Report\n")
    lines.append(f"**File:** `{metadata['path']}`  ")
    lines.append(f"**Size:** {metadata['size_mb']} MB  ")
    lines.append(f"**MD5:** `{metadata['md5']}`  ")
    lines.append(f"**Profiled at:** {metadata['profiled_at']}  ")
    lines.append(f"**Sheets:** {len(excel_file.sheet_names)}  ")
    lines.append("")

    # Hidden sheets warning
    if hidden_sheets:
        lines.append("### Hidden Sheets Detected\n")
        for hs in hidden_sheets:
            lines.append(f"- **{hs['name']}** (state: {hs['state']}) — review before including in pipeline")
        lines.append("")

    # Named ranges
    if named_ranges:
        lines.append("### Named Ranges\n")
        lines.append("| Name | Sheet | Range | Hidden |")
        lines.append("|------|-------|-------|--------|")
        for nr in named_ranges:
            lines.append(f"| {nr['name']} | {nr['sheet']} | {nr['range']} | {'yes' if nr['hidden'] else 'no'} |")
        lines.append("")
        lines.append("Note: Named ranges may define data regions, print areas, or validation lists.")
        lines.append("Hidden named ranges are often system-generated (e.g., filter ranges).\n")

    # Per-sheet profiles
    all_profiles = {}
    for sheet_name, df in sheets_data.items():
        lines.append(f"## Sheet: {sheet_name}\n")

        if len(df) < 2:
            lines.append(f"**Rows:** {len(df)} — SKIPPED (too few data rows)\n")
            continue

        lines.append(f"**Rows:** {len(df)} | **Columns:** {len(df.columns)}\n")

        # Formula warnings
        if sheet_name in formula_warnings:
            lines.append("### Formula Cell Warnings\n")
            for w in formula_warnings[sheet_name]:
                lines.append(f"- WARNING: {w}")
            lines.append("")

        # Merged cells
        if sheet_name in merged:
            lines.append("### Merged Cell Regions\n")
            lines.append("| Range | Rows | Cols | Value |")
            lines.append("|-------|------|------|-------|")
            for region in merged[sheet_name]:
                lines.append(f"| {region['range']} | {region['rows']} | {region['cols']} | {region['value'][:50]} |")
            lines.append("")

        # Transposition detection
        transposition = detect_transposition(df)
        if transposition["likely_transposed"]:
            lines.append("### Transposition Detected\n")
            lines.append(f"**Reason:** {transposition['reason']}  ")
            lines.append(f"**Suggestion:** {transposition['suggestion']}  ")
            lines.append("")

        # Column profiles
        lines.append("### Column Profiles\n")
        lines.append("| Column | Type | Nulls | Empty Str | Unique | Min/Max or Samples |")
        lines.append("|--------|------|-------|-----------|--------|-------------------|")
        for col in df.columns:
            p = profile_column(df[col])
            all_profiles[(sheet_name, col)] = p
            null_pct = f"{p['null_rate']*100:.1f}%"
            empty_str = str(p.get('empty_string_count', 0))
            unique = f"{p['unique_count']} ({p['unique_rate']*100:.0f}%)"

            detail = ""
            if "min" in p and "max" in p:
                detail = f"{p['min']} — {p['max']}"
            elif "samples" in p:
                detail = ", ".join(p["samples"][:3])
            elif "values" in p:
                detail = ", ".join(p["values"][:5])

            type_str = p["inferred_type"]
            if p.get("possible_enum"):
                type_str += " (enum?)"
            if p.get("mixed_type_warning"):
                type_str += " ⚠MIXED"
            if p.get("format_detected"):
                type_str += f" [{p['format_detected']}]"

            lines.append(f"| {col} | {type_str} | {null_pct} | {empty_str} | {unique} | {detail[:60]} |")
        lines.append("")

        # Mixed type warnings detail
        mixed_cols = [(col, p) for col, p in ((c, all_profiles.get((sheet_name, c), {})) for c in df.columns)
                      if p.get("mixed_type_warning")]
        if mixed_cols:
            lines.append("### Mixed Type Warnings\n")
            for col, p in mixed_cols:
                lines.append(f"- **{col}**: {p['mixed_type_warning']}")
                if p.get("coercion_failures"):
                    lines.append(f"  Sample non-numeric values: {p['coercion_failures']}")
            lines.append("")

        # Mojibake / encoding issues
        mojibake_cols = []
        for col in df.columns:
            if df[col].dtype == object:  # String columns only
                mj = detect_mojibake(df[col])
                if mj and mj["rate"] > 0.01:  # >1% of sampled values
                    mojibake_cols.append((col, mj))
        if mojibake_cols:
            lines.append("### Encoding Issues Detected (Mojibake)\n")
            for col, mj in mojibake_cols:
                lines.append(f"- **{col}**: {mj['affected_count']}/{mj['sample_size']} sampled values contain garbled characters")
                lines.append(f"  Examples: {mj['examples']}")
            lines.append("\n**Action:** The source file may have been saved with incorrect encoding. "
                         "Try re-exporting from the original system with UTF-8 encoding.\n")

        # Calculated columns
        calc = detect_calculated_columns(df)
        if calc:
            lines.append("### Calculated Columns Detected\n")
            lines.append("| Column | Formula | Match Rate | Mismatches |")
            lines.append("|--------|---------|------------|------------|")
            for c in calc:
                lines.append(f"| {c['column']} | `{c['formula']}` | {c['match_rate']*100:.1f}% | {c['mismatches']} |")
            # Show mismatch examples for diagnosis
            for c in calc:
                if c.get("mismatch_examples"):
                    lines.append(f"\n**{c['column']} mismatch examples** (for diagnosis):")
                    for ex in c["mismatch_examples"]:
                        lines.append(f"  Row {ex['row']}: got {ex[c['column']]}, expected {ex['expected']}, diff={ex['diff']:.6f}")
            lines.append("")

    # Cross-sheet FK candidates
    if fk_candidates:
        lines.append("## Cross-Sheet Relationships\n")
        lines.append("| Direction | Match Rate | Orphans | Name Match | Verification |")
        lines.append("|-----------|-----------|---------|------------|-------------|")
        for fk in fk_candidates:
            rate = fk['match_rate']
            if rate >= 0.99:
                verification = "VERIFIED"
            elif rate >= 0.95:
                verification = "VERIFIED (with warnings)"
            elif rate >= 0.80:
                verification = "POSSIBLE (review needed)"
            else:
                verification = "UNLIKELY"
            name_match = "yes" if fk.get("name_similarity") else "no"
            lines.append(f"| {fk['direction']} | {rate*100:.1f}% | {fk['orphan_count']} | {name_match} | {verification} |")
        lines.append("")

    # Flagged unknowns
    lines.append("## Flagged Unknowns\n")
    lines.append("The following require human knowledge or cannot be auto-verified:\n")
    for sheet_name, df in sheets_data.items():
        for col in df.columns:
            p = all_profiles.get((sheet_name, col), {})
            if p.get("possible_enum") and p.get("inferred_type") == "string":
                values = ", ".join(p.get("values", [])[:10])
                lines.append(f"- **{sheet_name}.{col}**: Enum-like values `[{values}]` — meanings unknown")
            if p.get("empty_string_count", 0) > 0:
                lines.append(f"- **{sheet_name}.{col}**: {p['empty_string_count']} empty strings — treat as NULL? (add NULLIF in staging)")
    if hidden_sheets:
        for hs in hidden_sheets:
            lines.append(f"- **{hs['name']}**: Hidden sheet — include in pipeline or skip?")
    lines.append("")

    report = "\n".join(lines)
    output.write_text(report)
    print(f"Report written to {output} ({len(lines)} lines)")


def main():
    parser = argparse.ArgumentParser(description="Profile an Excel file")
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--output", default="data-profile-report.md", help="Output report path")
    args = parser.parse_args()

    filepath = Path(args.excel_file)
    if not filepath.exists():
        print(f"Error: {filepath} not found")
        sys.exit(1)

    generate_report(filepath, Path(args.output))


if __name__ == "__main__":
    main()
